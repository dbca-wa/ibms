from django.db.models import Sum
from openpyxl.styles import Alignment, Border, Side

from ibms.models import IBMData, GLPivDownload, SFMServicePriority
from sfm.models import MeasurementValue


def outputs_report(workbook, sfm, spn, dp, fy, qtr, cc):
    sheet = workbook.active
    # Insert values in header area.
    sheet['B2'] = qtr.description
    sheet['C2'] = fy.financialYear
    sheet['B3'] = cc.costCentre
    sheet['E3'] = dp
    curr = '"$"#,##0_);("$"#,##'  # An Excel currency format.
    wrapped = Alignment(vertical='top', wrap_text=True)
    row = 6  # Note that openpyxl start row indexing at 1.

    # For each of the service priority numbers, take a subset of the query
    # and insert values as required.
    for s in spn:
        # Where there are no SFMServicePriority objects for the given
        # servicePriorityNo and financialYear, skip the priority no.
        if not SFMServicePriority.objects.filter(
                servicePriorityNo=s, financialYear=fy.financialYear).exists():
            continue
        sfm_metrics = sfm.filter(servicePriorityNo=s)
        sfm_service_pri = SFMServicePriority.objects.get(
            servicePriorityNo=s, financialYear=fy.financialYear)
        ibm_data = IBMData.objects.filter(
            financialYear=fy.financialYear, servicePriorityID=s,
            costCentre=cc.costCentre)
        ibm_ids = set(ibm_data.values_list('ibmIdentifier', flat=True))
        gl = GLPivDownload.objects.filter(
            financialYear=fy.financialYear, costCentre=cc.costCentre,
            codeID__in=ibm_ids)
        for k, metric in enumerate(sfm_metrics):  # Subquery
            if k == 0:  # First row only of subquery.
                cell = sheet.cell(column=1, row=row, value=sfm_service_pri.servicePriorityNo)
                cell.alignment = wrapped
                cell = sheet.cell(column=2, row=row, value=sfm_service_pri.description)
                cell.alignment = wrapped
                cell = sheet.cell(column=3, row=row, value=sfm_service_pri.description2)
                cell.alignment = wrapped
                ytd_a = gl.aggregate(Sum('ytdActual'))
                cell = sheet.cell(column=4, row=row, value=ytd_a['ytdActual__sum'])
                cell.number_format = curr
                cell.alignment = wrapped
                ytd_b = gl.aggregate(Sum('ytdBudget'))
                cell = sheet.cell(column=5, row=row, value=ytd_b['ytdBudget__sum'])
                cell.number_format = curr
                cell.alignment = wrapped
            cell = sheet.cell(column=6, row=row, value=metric.metricID)
            cell.alignment = wrapped
            cell = sheet.cell(column=7, row=row, value=metric.descriptor)
            cell.alignment = wrapped
            measure_type_col = {
                'Quantity': 8,
                'Percentage': 9,
                'Hectare': 10,
                'Kilometer': 11}
            ytd_measure = {
                'Quantity': 0,
                'Percentage': 0,
                'Hectare': 0,
                'Kilometer': 0}
            quarter_col = {
                'Q1 (Jul - Sep)': 16,
                'Q2 (Oct - Dec)': 17,
                'Q3 (Jan - Mar)': 18,
                'Q4 (Apr - Jun)': 19}
            measurements_ytd = MeasurementValue.objects.filter(
                quarter__financialYear=qtr.financialYear,
                costCentre=cc, sfmMetric=metric, value__isnull=False)
            for m in measurements_ytd:
                # Aggregate the YTD measurements of each unit type:
                ytd_measure[m.measurementType.unit] += m.value
                # Write all quarter achievements text (columns 12-15)
                cell = sheet.cell(column=quarter_col[m.quarter.description], row=row, value=m.comment)
                cell.alignment = wrapped
            # Write the YTD measurements (columns 7-10):
            for k, v in ytd_measure.items():
                if v > 0:
                    cell = sheet.cell(column=measure_type_col[k], row=row, value=v)
                    cell.alignment = wrapped
            # Current quarter measurements (columns 11-14)
            measurements_qtr = MeasurementValue.objects.filter(
                quarter=qtr, costCentre=cc, sfmMetric=metric, value__isnull=False)
            if measurements_qtr:
                for m in measurements_qtr:
                    cell = sheet.cell(column=measure_type_col[m.measurementType.unit] + 4, row=row, value=m.value)
                    cell.alignment = wrapped

            # Set the row height to 16.5 (22 pixels).
            sheet.row_dimensions[row].ht = 16.5

            row += 1

    # Re-apply cell borders in the header row.
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    for c in [sheet.cell(column=col, row=5) for col in range(1, 20)]:
        c.border = thin_border
